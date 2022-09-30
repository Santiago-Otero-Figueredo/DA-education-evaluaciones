from django.core.management.base import BaseCommand
from django.contrib.auth.models import Group

from apps.usuarios.models import Usuario
from apps.matriculas.models import Matricula
from apps.periodos_academicos.models import PeriodoAcademico
from apps.cursos.models import Curso, CursoDelPrograma
from apps.competencias.models import ProfesorCursoPrograma, TipoNivelEvaluacion, NivelEvaluacion, Grupo
from apps.programas.models import Programa
from apps.evaluaciones.models import Actividad, Nota, Pregunta, RestriccionNivelProfesorCurso, TipoActividad, CalificacionEvaluacion, RestriccionNivelActividad, CalificacionCualitativa

import names

import openpyxl
from django.conf import settings

UBICACION_ARCHIVOS ="/_datos_iniciales/produccion/"

class Command(BaseCommand):
    help = 'Guarda un archivo de backup.'

    def handle(self, *args, **kwargs):
        # cargar_grupos(self)
        # cargar_programas(self)
        # cargar_cursos(self)
        # cargar_cursos_del_programa(self)
        # cargar_estudiantes(self)
        # cargar_profesores(self)
        # cargar_periodos_academicos(self)
        # cargar_matriculas(self)
        # cargar_profesor_curso_programa(self)
        # cargar_tipos_nivel_evaluacion(self)
        # cargar_niveles_evaluacion(self)
        # cargar_restriccion_nivel_profesor(self)
        # cargar_tipos_actividad(self)
        # cargar_restricciones_nivel_actividad(self)
        # cargar_evaluaciones(self)
        # cargar_preguntas(self)
        # cargar_calificaciones_cualitativas(self)
        # cargar_grupos_profesor(self)
        cargar_grupos_estudiantes(self)
        # cargar_notas(self)
        # generar_notas_automaticamente(self)

        pass

def cargar_grupos(handle):

    if len(Group.objects.all()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}01_grupos.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            name = str(worksheet[i][1].value)
            Group.objects.create(
                id=id,
                name=name
            )

        handle.stdout.write(handle.style.SUCCESS("Calificaciones evaluaciones creados"))

def cargar_programas(handle):

    if len(Programa.obtener_todos()) == 0:

        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}02_programas.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            nombre = str(worksheet[i][2].value)
            codigo = str(int(worksheet[i][3].value))
            nivel_academico = str(worksheet[i][4].value)

            Programa.objects.create(
                id=id,
                nombre=nombre,
                codigo=codigo,
                nivel_academico=nivel_academico,
                activa=activa
            )

        handle.stdout.write(handle.style.SUCCESS("Programas creados"))

def cargar_cursos(handle):

    if len(Curso.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}03_cursos.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            nombre = str(worksheet[i][2].value)
            codigo = str(worksheet[i][3].value)

            Curso.objects.create(
                id=id,
                nombre=nombre,
                codigo=codigo,
                activa=activa
            )

        handle.stdout.write(handle.style.SUCCESS("Cursos creados"))

def cargar_cursos_del_programa(handle):

    if len(CursoDelPrograma.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}04_cursos_del_programa.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            curso = int(worksheet[i][2].value)
            if worksheet[i][3].value == None:
                prerrequisito = None
            else:
                prerrequisito = int(worksheet[i][3].value)
            programa = int(worksheet[i][4].value)

            CursoDelPrograma.objects.create(
                id=id,
                curso=Curso.obtener_por_id(curso),
                programa=Programa.obtener_por_id(programa),
                prerrequisito=CursoDelPrograma.obtener_por_id(prerrequisito),
                activa=activa
            )
        handle.stdout.write(handle.style.SUCCESS("Curso del programa creados"))

def cargar_estudiantes(handle):

    if len(Usuario.obtener_estudiantes()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}05_usuarios.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            password = str(worksheet[i][1].value)
            username = str(worksheet[i][4].value)
            nombre = str(worksheet[i][5].value)
            apellido = str(worksheet[i][6].value)
            email = str(worksheet[i][7].value)
            codigo = str(int(worksheet[i][12].value))

            usuario = Usuario.objects.create(
                id=id,
                username=username,
                first_name=nombre,
                last_name=apellido,
                email=email,
                password=password,
                codigo=codigo
            )

            rol = Group.objects.get(name='Estudiante')
            rol.user_set.add(usuario)
        handle.stdout.write(handle.style.SUCCESS("Usuarios creados"))

def cargar_profesores(handle):

    if len(Usuario.obtener_profesores()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}05.2_profesores.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            password = str(worksheet[i][1].value)
            username = str(worksheet[i][4].value)
            nombre = str(worksheet[i][5].value)
            apellido = str(worksheet[i][6].value)
            email = str(worksheet[i][7].value)
            codigo = str(int(worksheet[i][12].value))

            usuario = Usuario.objects.create(
                id=id,
                username=username,
                first_name=nombre,
                last_name=apellido,
                email=email,
                password=password,
                codigo=codigo
            )

            rol = Group.objects.get(name='Profesor')
            rol.user_set.add(usuario)

        handle.stdout.write(handle.style.SUCCESS("Profesores creados"))

def cargar_periodos_academicos(handle):

    if len(PeriodoAcademico.obtener_activos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}06_periodos_academicos.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            ano = int(worksheet[i][2].value)
            periodo = str(worksheet[i][3].value)
            fecha_inicio = str(worksheet[i][4].value).split(" ")[0]
            fecha_fin = str(worksheet[i][5].value).split(" ")[0]
            PeriodoAcademico.objects.create(
                id=id,
                ano=ano,
                periodo=periodo,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                activa=activa
            )

        handle.stdout.write(handle.style.SUCCESS("Periodos académicos creados"))

def cargar_matriculas(handle):

    if not Matricula.objects.filter(periodo_academico=1).exists():

        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}07_matriculas.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            aprobada = str(worksheet[i][2].value)=='t'
            curso_del_programa = int(worksheet[i][3].value)
            estudiante = int(worksheet[i][4].value)
            periodo_academico = int(worksheet[i][5].value)

            Matricula.objects.create(
                id=id,
                aprobada=aprobada,
                activa=activa,
                curso_del_programa=CursoDelPrograma.obtener_por_id(curso_del_programa),
                periodo_academico=PeriodoAcademico.obtener_por_id(periodo_academico),
                estudiante=Usuario.obtener_por_id(estudiante)
            )

        handle.stdout.write(handle.style.SUCCESS("Matriculas creadas"))

def cargar_profesor_curso_programa(handle):

    if len(ProfesorCursoPrograma.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}08_profesor_cursos_programa.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            curso_programa = int(worksheet[i][2].value)
            periodo_academico = int(worksheet[i][3].value)
            profesor = int(worksheet[i][4].value)

            ProfesorCursoPrograma.objects.create(
                id=id,
                curso_programa=CursoDelPrograma.obtener_por_id(curso_programa),
                periodo_academico=PeriodoAcademico.obtener_por_id(periodo_academico),
                profesor=Usuario.obtener_por_id(profesor),
                activa=activa
            )


        handle.stdout.write(handle.style.SUCCESS("Profesores curso programa creados"))

def cargar_tipos_nivel_evaluacion(handle):

    if len(TipoNivelEvaluacion.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}09_tipos_nivel_evaluacion.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            nombre = str(worksheet[i][2].value)
            descripcion = str(worksheet[i][3].value)
            nivel = int(worksheet[i][4].value)
            mide_programa = str(worksheet[i][5].value)=='t'
            if worksheet[i][6].value == None:
                tipo_superior = None
            else:
                tipo_superior = TipoNivelEvaluacion.obtener_por_id(int(worksheet[i][6].value))

            TipoNivelEvaluacion.objects.create(
                id=id,
                nombre=nombre,
                descripcion=descripcion,
                mide_programa=mide_programa,
                tipo_superior=tipo_superior,
                nivel=nivel,
                activa=activa
            )


        handle.stdout.write(handle.style.SUCCESS("Tipos nivel evaluación creados"))

def cargar_niveles_evaluacion(handle):

    if not NivelEvaluacion.objects.filter(curso_del_programa__pk=1).exists():

        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}10_niveles_evaluacion.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            nombre = str(worksheet[i][2].value)
            descripcion = str(worksheet[i][3].value)
            mide_porcentaje = str(worksheet[i][4].value)=='t'
            porcentaje = float(worksheet[i][5].value)
            curso_del_programa = int(worksheet[i][6].value)
            if worksheet[i][7].value == None:
                nivel_asociado = None
            else:
                nivel_asociado = NivelEvaluacion.obtener_por_id(int(worksheet[i][7].value))
            tipo_nivel = int(worksheet[i][8].value)

            NivelEvaluacion.objects.create(
                id=id,
                nombre=nombre,
                descripcion=descripcion,
                mide_porcentaje=mide_porcentaje,
                porcentaje=porcentaje,
                tipo_nivel=TipoNivelEvaluacion.obtener_por_id(tipo_nivel),
                nivel_asociado=nivel_asociado,
                curso_del_programa=CursoDelPrograma.obtener_por_id(curso_del_programa),
                activa=activa
            )

        handle.stdout.write(handle.style.SUCCESS("Niveles de evaluación creados"))

def cargar_restriccion_nivel_profesor(handle):

    if len(RestriccionNivelProfesorCurso.objects.filter(profesor_curso_programa__pk=1)) == 0:

        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}11_restriccion_nivel_profesor_curso.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            porcentaje = float(worksheet[i][1].value)
            nivel_evaluacion = int(worksheet[i][2].value)
            profesor_curso_programa = int(worksheet[i][3].value)

            RestriccionNivelProfesorCurso.objects.create(
                id=id,
                porcentaje=porcentaje,
                nivel_evaluacion=NivelEvaluacion.obtener_por_id(nivel_evaluacion),
                profesor_curso_programa=ProfesorCursoPrograma.obtener_por_id(profesor_curso_programa),
                activa=True
            )

        handle.stdout.write(handle.style.SUCCESS("Calificaciones nivel evaluacion creados"))

def cargar_tipos_actividad(handle):

    if len(TipoActividad.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}12_tipos_actividad.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            nombre = str(worksheet[i][2].value)

            TipoActividad.objects.create(
                id=id,
                nombre=nombre,
                activa=activa
            )

        handle.stdout.write(handle.style.SUCCESS("Actividades evaluacion creados"))

def cargar_restricciones_nivel_actividad(handle):

    if len(RestriccionNivelActividad.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}13_restricciones_nivel_actividad.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            tipo_actividad = int(worksheet[i][1].value)
            restriccion_nivel_profesor_curso = int(worksheet[i][2].value)
            porcentaje = float(worksheet[i][3].value)

            RestriccionNivelActividad.objects.create(
                id=id,
                tipo_actividad=TipoActividad.obtener_por_id(tipo_actividad),
                restriccion_nivel_profesor_curso=RestriccionNivelProfesorCurso.obtener_por_id(restriccion_nivel_profesor_curso),
                porcentaje=porcentaje
            )

        handle.stdout.write(handle.style.SUCCESS("Restricciones nivel actividad creados"))

def cargar_evaluaciones(handle):

    if len(Actividad.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}14_actividades.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            profesor_curso_programa = int(worksheet[i][1].value)
            tipo_calificacion = str(worksheet[i][2].value)
            tipo_actividad = int(worksheet[i][3].value)
            nombre = str(worksheet[i][4].value)

            Actividad.objects.create(
                id=id,
                profesor_curso_programa=ProfesorCursoPrograma.obtener_por_id(profesor_curso_programa),
                tipo_actividad=TipoActividad.obtener_por_id(tipo_actividad),
                tipo_calificacion=tipo_calificacion,
                nombre=nombre
            )


        handle.stdout.write(handle.style.SUCCESS("Actividades creadas"))

def cargar_preguntas(handle):

    if len(Pregunta.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}15_preguntas.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            actividad = int(worksheet[i][1].value)
            restriccion_nivel_actividad = int(worksheet[i][2].value)
            numeracion = int(worksheet[i][3].value)
            contenido = str(worksheet[i][4].value)
            porcentaje = float(worksheet[i][5].value)

            Pregunta.objects.create(
                id=id,
                actividad=Actividad.obtener_por_id(actividad),
                restriccion_nivel_actividad=RestriccionNivelActividad.obtener_por_id(restriccion_nivel_actividad),
                numeracion=numeracion,
                contenido=contenido,
                porcentaje=porcentaje
            )


        handle.stdout.write(handle.style.SUCCESS("Preguntas creadas"))

def cargar_calificaciones_cualitativas(handle):

    if len(CalificacionCualitativa.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}16_calificacion_cualtativa.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            nombre = str(worksheet[i][1].value)
            valor_minimo = float(worksheet[i][2].value)
            valor_maximo = float(worksheet[i][3].value)


            CalificacionCualitativa.objects.create(
                id=id,
                nombre=nombre,
                valor_minimo=valor_minimo,
                valor_maximo=valor_maximo
            )


        handle.stdout.write(handle.style.SUCCESS("Calificaciones cualitativas creadas"))

def cargar_grupos_profesor(handle):

    if len(Grupo.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}17_grupos_profesor.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            profesor_curso_programa = int(worksheet[i][1].value)
            nombre = str(worksheet[i][2].value)

            Grupo.objects.create(
                id=id,
                profesor_curso_programa=ProfesorCursoPrograma.obtener_por_id(profesor_curso_programa),
                nombre=nombre,
            )

        handle.stdout.write(handle.style.SUCCESS("Grupos profesor creados"))

def cargar_grupos_estudiantes(handle):

    workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}18_grupos_estudiante.xlsx')
    worksheet = workbook.active

    for i in range(2, worksheet.max_row + 1):
        id = int(worksheet[i][0].value)
        grupo = int(worksheet[i][1].value)
        estudiante = int(worksheet[i][2].value)

        grupo = Grupo.obtener_por_id(grupo)
        estudiante = Usuario.obtener_por_id(estudiante)
        grupo.estudiantes.add(estudiante)

    handle.stdout.write(handle.style.SUCCESS("Grupos estudiantes creados"))

def cargar_notas(handle):

    if len(Nota.obtener_todos()) == 0:
        workbook = openpyxl.load_workbook(f'{settings.ROOT_DIR}{UBICACION_ARCHIVOS}19_notas.xlsx')
        worksheet = workbook.active

        for i in range(2, worksheet.max_row + 1):
            id = int(worksheet[i][0].value)
            activa = str(worksheet[i][1].value)=='t'
            resultado = float(worksheet[i][2].value)
            if worksheet[i][3].value == None:
                calificacion_cualitativa = None
            else:
                calificacion_cualitativa = CalificacionCualitativa.obtener_por_id(int(worksheet[i][3].value))
            estudiante = int(worksheet[i][4].value)
            pregunta = int(worksheet[i][5].value)

            Nota.objects.create(
                id=id,
                calificacion_cualitativa=calificacion_cualitativa,
                pregunta=Pregunta.obtener_por_id(pregunta),
                estudiante=Usuario.obtener_por_id(estudiante),
                resultado=resultado,
                activa=activa
            )

        handle.stdout.write(handle.style.SUCCESS("Notas creadas"))

def generar_notas_automaticamente(handle):
    import random

    if len(Nota.obtener_todos()) == 0:
        calificacion_cualitativa = None

        for profesor_curso_programa in ProfesorCursoPrograma.obtener_todos():
            estudiantes = set(list(Usuario.obtener_estudiantes().filter(
                profesor_curso_programa_asociados_a_estudiante__pk=profesor_curso_programa.pk
            ).values_list('pk', flat=True)))
            # estudiantes = set(list(Usuario.obtener_estudiantes().filter(
            #     matriculas_del_estudiante__periodo_academico__pk=profesor_curso_programa.periodo_academico.pk,
            #     matriculas_del_estudiante__curso_del_programa__profesor_curso_programa_asociados__pk=profesor_curso_programa.pk,
            # ).values_list('pk', flat=True)))

            # obtener preguntas por curso del programa
            preguntas = Pregunta.objects.filter(
                restriccion_nivel_actividad__restriccion_nivel_profesor_curso__profesor_curso_programa__pk=profesor_curso_programa.pk
            )

            preguntas_id = list(set(preguntas.values_list('pk', flat=True)))
            for pregunta in Pregunta.obtener_por_lista_ids(preguntas_id):
                estudiantes_objetos = Usuario.objects.filter(pk__in=estudiantes)


                for estudiante_objeto in estudiantes_objetos:

                    calificacion_cualitativa = None
                    resultado = random.randint(100, 500) / 100
                    if pregunta.actividad.tipo_calificacion == "Cualitativa":
                        calificacion_cualitativa = CalificacionCualitativa.obtener_calificacion_por_numero(resultado)
                        if calificacion_cualitativa is None:
                            calificacion_cualitativa = CalificacionCualitativa.objects.get(nombre='Deficiente')

                        resultado = 0.0


                    Nota.objects.create(
                        calificacion_cualitativa=calificacion_cualitativa,
                        pregunta=pregunta,
                        estudiante=estudiante_objeto,
                        resultado=resultado
                    )
        handle.stdout.write(handle.style.SUCCESS("Notas creadas"))
